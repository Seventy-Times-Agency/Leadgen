from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from leadgen.db.models import Lead


COLUMNS: list[tuple[str, int]] = [
    ("Название", 36),
    ("AI-скор", 10),
    ("Теги", 18),
    ("Резюме", 40),
    ("Совет: как зайти", 50),
    ("Сильные стороны", 35),
    ("Точки роста / слабые", 35),
    ("Риски", 30),
    ("Категория", 22),
    ("Телефон", 18),
    ("Сайт", 32),
    ("Соцсети", 24),
    ("Адрес", 45),
    ("Рейтинг Google", 12),
    ("Отзывов", 10),
    ("Отзывы (кратко)", 45),
    ("Широта", 12),
    ("Долгота", 12),
    ("Источник", 14),
]

HEADER_FILL = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")


def build_excel(leads: Iterable[Lead]) -> bytes:
    """Render a list of leads into an XLSX file and return its bytes."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Leads"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    body_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, lead in enumerate(leads, start=2):
        ws.cell(row=row_idx, column=1, value=lead.name)
        ws.cell(
            row=row_idx,
            column=2,
            value=int(lead.score_ai) if lead.score_ai is not None else None,
        )
        ws.cell(row=row_idx, column=3, value=", ".join(lead.tags or []))
        ws.cell(row=row_idx, column=4, value=lead.summary)
        ws.cell(row=row_idx, column=5, value=lead.advice)
        ws.cell(row=row_idx, column=6, value="\n".join(lead.strengths or []))
        ws.cell(row=row_idx, column=7, value="\n".join(lead.weaknesses or []))
        ws.cell(row=row_idx, column=8, value="\n".join(lead.red_flags or []))
        ws.cell(row=row_idx, column=9, value=lead.category)
        ws.cell(row=row_idx, column=10, value=lead.phone)
        ws.cell(row=row_idx, column=11, value=lead.website)
        ws.cell(
            row=row_idx,
            column=12,
            value=", ".join((lead.social_links or {}).keys()),
        )
        ws.cell(row=row_idx, column=13, value=lead.address)
        ws.cell(row=row_idx, column=14, value=lead.rating)
        ws.cell(row=row_idx, column=15, value=lead.reviews_count)
        ws.cell(row=row_idx, column=16, value=lead.reviews_summary)
        ws.cell(row=row_idx, column=17, value=lead.latitude)
        ws.cell(row=row_idx, column=18, value=lead.longitude)
        ws.cell(row=row_idx, column=19, value=lead.source)

        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = body_align

    ws.freeze_panes = "C2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
